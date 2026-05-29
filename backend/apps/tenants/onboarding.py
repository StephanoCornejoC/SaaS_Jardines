"""
Importador de jardín nuevo desde la plantilla Excel de Miniddo.

Flujo expuesto al admin Django:
  1. `parse_excel(file_obj)` — lee el .xlsx a un dict y devuelve errores de
     formato (DNIs malformados, fechas inválidas, enums fuera de lista, etc.).
  2. `validate_cross_sheet(data)` — reglas de negocio entre hojas (FKs por
     nombre/DNI, principal único por alumno, tipos coherentes, etc.).
  3. `is_schema_empty(tenant)` — sanity check antes de ejecutar: el import
     asume jardín VACÍO (decisión #2 del onboarding).
  4. `execute_import(tenant, data, user)` — crea TODO dentro de un
     `transaction.atomic()` en el schema del tenant. Si algo falla a mitad,
     rollback completo y el jardín queda intacto.

El admin orquesta los pasos en una vista de 2 fases (preview + confirmar).
"""
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from django.db import transaction
from django_tenants.utils import schema_context


# ---------------------------------------------------------------------------
# Constantes del dominio
# ---------------------------------------------------------------------------

EXPECTED_SHEETS = (
    "1_Aulas", "2_Profesores", "3_Alumnos", "4_Apoderados", "5_Configuracion",
)

MES_INICIO_PENSION = 3
MES_FIN_PENSION = 12
DIA_VENCIMIENTO_DEFAULT = 15

VALID_NIVELES = {2, 3, 4, 5}
VALID_TIPOS_PROFESOR = {"TITULAR", "AUXILIAR"}
VALID_GENEROS = {"M", "F"}
VALID_PARENTESCOS = {"PADRE", "MADRE", "TUTOR", "OTRO"}
VALID_SI_NO = {"SI", "NO"}

ANIO_MIN = 2020
ANIO_MAX = 2100
EDAD_MIN_ALUMNO = 1
EDAD_MAX_ALUMNO = 6


# ---------------------------------------------------------------------------
# Helpers de parsing por celda
# ---------------------------------------------------------------------------

def _err(hoja: str, fila: Any, msg: str) -> dict:
    return {"hoja": hoja, "fila": fila, "msg": msg}


def _norm_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_dni(value, hoja, fila):
    s = _norm_str(value)
    if not s:
        return None, _err(hoja, fila, "DNI vacío")
    # Excel a veces convierte DNIs con leading zeros a int (47000001 → 47000001
    # OK, pero 04700000 → 4700000). Re-paddear a 8 chars.
    s = s.zfill(8)
    if not re.fullmatch(r"\d{8}", s):
        return None, _err(hoja, fila, f"DNI inválido: {value!r} (deben ser 8 dígitos)")
    return s, None


def _parse_fecha(value, hoja, fila, campo):
    if value is None or _norm_str(value) == "":
        return None, _err(hoja, fila, f"{campo}: vacío")
    # openpyxl puede devolver datetime cuando la celda tiene formato fecha
    # de Excel. Normalizamos a date para comparaciones uniformes.
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    try:
        parts = _norm_str(value).replace("-", "/").split("/")
        if len(parts) != 3:
            raise ValueError
        d, m, y = (int(p) for p in parts)
        return date(y, m, d), None
    except Exception:
        return None, _err(hoja, fila, f"{campo}: fecha inválida {value!r} (use DD/MM/AAAA)")


def _parse_int(value, hoja, fila, campo, minimo=None, maximo=None):
    if value is None or _norm_str(value) == "":
        return None, _err(hoja, fila, f"{campo}: vacío")
    try:
        n = int(value)
    except (TypeError, ValueError):
        return None, _err(hoja, fila, f"{campo}: número inválido {value!r}")
    if minimo is not None and n < minimo:
        return None, _err(hoja, fila, f"{campo}: debe ser ≥ {minimo}")
    if maximo is not None and n > maximo:
        return None, _err(hoja, fila, f"{campo}: debe ser ≤ {maximo}")
    return n, None


def _parse_decimal(value, hoja, fila, campo, *, required=True):
    if value is None or _norm_str(value) == "":
        if required:
            return None, _err(hoja, fila, f"{campo}: vacío")
        return None, None
    try:
        d = Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None, _err(hoja, fila, f"{campo}: monto inválido {value!r}")
    if d < 0:
        return None, _err(hoja, fila, f"{campo}: no puede ser negativo")
    return d, None


def _parse_enum(value, valid_set, hoja, fila, campo, *, required=True):
    s = _norm_str(value).upper()
    if not s:
        if required:
            return None, _err(hoja, fila, f"{campo}: vacío")
        return None, None
    if s not in valid_set:
        return None, _err(hoja, fila,
                          f"{campo}: '{value}' no permitido. Use: {', '.join(sorted(valid_set))}")
    return s, None


def _calcular_edad(fecha_nac, ref=None):
    ref = ref or date.today()
    return ref.year - fecha_nac.year - ((ref.month, ref.day) < (fecha_nac.month, fecha_nac.day))


# ---------------------------------------------------------------------------
# Parser por hoja
# ---------------------------------------------------------------------------

def _iter_data_rows(ws, header_row=1):
    """Itera filas desde la 2 en adelante, saltando filas vacías y la fila
    de ejemplo en gris (heurística: si la primera celda contiene '(ejemplo)'
    en minúsculas, la skippeamos por respeto a la fila de muestra)."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                  start=header_row + 1):
        if all((c is None or _norm_str(c) == "") for c in row):
            continue
        # Skip fila de ejemplo si trae el marcador "(ejemplo)" en alguna celda
        first_text = " ".join(_norm_str(c).lower() for c in row if isinstance(c, str))
        if "(ejemplo)" in first_text:
            continue
        yield row_idx, row


def _parse_aulas(ws, errors):
    """Cols: nombre, nivel_edad, capacidad"""
    result = []
    for row_idx, row in _iter_data_rows(ws):
        nombre = _norm_str(row[0] if len(row) > 0 else None)
        if not nombre:
            errors.append(_err("1_Aulas", row_idx, "nombre: vacío"))
            continue
        nivel, e = _parse_int(row[1] if len(row) > 1 else None, "1_Aulas", row_idx,
                              "nivel_edad", minimo=2, maximo=5)
        if e:
            errors.append(e); continue
        if nivel not in VALID_NIVELES:
            errors.append(_err("1_Aulas", row_idx, f"nivel_edad debe ser 2, 3, 4 o 5"))
            continue
        capacidad, e = _parse_int(row[2] if len(row) > 2 else None, "1_Aulas", row_idx,
                                  "capacidad", minimo=1, maximo=200)
        if e:
            errors.append(e); continue
        result.append({
            "_fila": row_idx,
            "nombre": nombre,
            "nivel_edad": nivel,
            "capacidad": capacidad,
        })
    return result


def _parse_profesores(ws, errors):
    """Cols: dni, nombres, apellidos, tipo, telefono, email, especialidad,
    fecha_ingreso, aula_titular, aula_auxiliar"""
    result = []
    for row_idx, row in _iter_data_rows(ws):
        dni, e = _parse_dni(row[0] if len(row) > 0 else None, "2_Profesores", row_idx)
        if e:
            errors.append(e); continue
        nombres = _norm_str(row[1] if len(row) > 1 else None)
        apellidos = _norm_str(row[2] if len(row) > 2 else None)
        if not nombres or not apellidos:
            errors.append(_err("2_Profesores", row_idx, "nombres y apellidos son obligatorios"))
            continue
        tipo, e = _parse_enum(row[3] if len(row) > 3 else None, VALID_TIPOS_PROFESOR,
                              "2_Profesores", row_idx, "tipo")
        if e:
            errors.append(e); continue
        telefono = _norm_str(row[4] if len(row) > 4 else None)
        if not telefono:
            errors.append(_err("2_Profesores", row_idx, "telefono: vacío"))
            continue
        email = _norm_str(row[5] if len(row) > 5 else None)
        especialidad = _norm_str(row[6] if len(row) > 6 else None)
        fecha_ingreso, e = _parse_fecha(row[7] if len(row) > 7 else None,
                                        "2_Profesores", row_idx, "fecha_ingreso")
        if e:
            errors.append(e); continue
        aula_titular = _norm_str(row[8] if len(row) > 8 else None) or None
        aula_auxiliar = _norm_str(row[9] if len(row) > 9 else None) or None

        result.append({
            "_fila": row_idx,
            "dni": dni, "nombres": nombres, "apellidos": apellidos,
            "tipo": tipo, "telefono": telefono, "email": email,
            "especialidad": especialidad, "fecha_ingreso": fecha_ingreso,
            "aula_titular": aula_titular, "aula_auxiliar": aula_auxiliar,
        })
    return result


def _parse_alumnos(ws, errors):
    """Cols: dni, nombres, apellidos, fecha_nacimiento, genero, aula,
    monto_matricula (opcional), monto_pension (opcional)"""
    result = []
    for row_idx, row in _iter_data_rows(ws):
        dni, e = _parse_dni(row[0] if len(row) > 0 else None, "3_Alumnos", row_idx)
        if e:
            errors.append(e); continue
        nombres = _norm_str(row[1] if len(row) > 1 else None)
        apellidos = _norm_str(row[2] if len(row) > 2 else None)
        if not nombres or not apellidos:
            errors.append(_err("3_Alumnos", row_idx, "nombres y apellidos son obligatorios"))
            continue
        fecha_nac, e = _parse_fecha(row[3] if len(row) > 3 else None,
                                    "3_Alumnos", row_idx, "fecha_nacimiento")
        if e:
            errors.append(e); continue
        if fecha_nac > date.today():
            errors.append(_err("3_Alumnos", row_idx, "fecha_nacimiento: no puede ser futura"))
            continue
        edad = _calcular_edad(fecha_nac)
        if edad < EDAD_MIN_ALUMNO or edad > EDAD_MAX_ALUMNO:
            errors.append(_err("3_Alumnos", row_idx,
                               f"edad {edad} fuera de rango ({EDAD_MIN_ALUMNO}-{EDAD_MAX_ALUMNO} años)"))
            continue
        genero, e = _parse_enum(row[4] if len(row) > 4 else None, VALID_GENEROS,
                                "3_Alumnos", row_idx, "genero")
        if e:
            errors.append(e); continue
        aula = _norm_str(row[5] if len(row) > 5 else None)
        if not aula:
            errors.append(_err("3_Alumnos", row_idx, "aula: vacío"))
            continue
        monto_matricula, e = _parse_decimal(row[6] if len(row) > 6 else None,
                                            "3_Alumnos", row_idx, "monto_matricula",
                                            required=False)
        if e:
            errors.append(e); continue
        monto_pension, e = _parse_decimal(row[7] if len(row) > 7 else None,
                                          "3_Alumnos", row_idx, "monto_pension",
                                          required=False)
        if e:
            errors.append(e); continue

        result.append({
            "_fila": row_idx,
            "dni": dni, "nombres": nombres, "apellidos": apellidos,
            "fecha_nacimiento": fecha_nac, "genero": genero, "aula": aula,
            "monto_matricula": monto_matricula, "monto_pension": monto_pension,
        })
    return result


def _parse_apoderados(ws, errors):
    """Cols: dni_alumno, dni, nombres, apellidos, telefono, email,
    parentesco, es_principal"""
    result = []
    for row_idx, row in _iter_data_rows(ws):
        dni_alumno, e = _parse_dni(row[0] if len(row) > 0 else None,
                                   "4_Apoderados", row_idx)
        if e:
            errors.append({**e, "msg": e["msg"].replace("DNI", "dni_alumno")}); continue
        dni, e = _parse_dni(row[1] if len(row) > 1 else None, "4_Apoderados", row_idx)
        if e:
            errors.append(e); continue
        nombres = _norm_str(row[2] if len(row) > 2 else None)
        apellidos = _norm_str(row[3] if len(row) > 3 else None)
        if not nombres or not apellidos:
            errors.append(_err("4_Apoderados", row_idx, "nombres y apellidos son obligatorios"))
            continue
        telefono = _norm_str(row[4] if len(row) > 4 else None)
        if not telefono:
            errors.append(_err("4_Apoderados", row_idx, "telefono: vacío"))
            continue
        email = _norm_str(row[5] if len(row) > 5 else None)
        parentesco, e = _parse_enum(row[6] if len(row) > 6 else None, VALID_PARENTESCOS,
                                    "4_Apoderados", row_idx, "parentesco")
        if e:
            errors.append(e); continue
        sino, e = _parse_enum(row[7] if len(row) > 7 else None, VALID_SI_NO,
                              "4_Apoderados", row_idx, "es_principal")
        if e:
            errors.append(e); continue

        result.append({
            "_fila": row_idx,
            "dni_alumno": dni_alumno, "dni": dni, "nombres": nombres,
            "apellidos": apellidos, "telefono": telefono, "email": email,
            "parentesco": parentesco, "es_principal": sino == "SI",
        })
    return result


def _parse_config(ws, errors):
    """Hoja con formato (Campo, Valor) — fila 8 es header, 9+ son datos."""
    HOJA = "5_Configuracion"
    cfg = {}
    # Buscar valores por nombre de campo en columna B, valor en columna C
    for row in ws.iter_rows(min_row=2, max_row=30, values_only=False):
        if not row or len(row) < 3:
            continue
        campo_cell = row[1]  # B
        valor_cell = row[2]  # C
        if campo_cell.value is None:
            continue
        campo = _norm_str(campo_cell.value).rstrip("*").strip()
        if campo == "anio_escolar":
            v, e = _parse_int(valor_cell.value, HOJA, campo_cell.row,
                              "anio_escolar", minimo=ANIO_MIN, maximo=ANIO_MAX)
            if e:
                errors.append(e)
            else:
                cfg["anio_escolar"] = v
        elif campo == "monto_matricula_base":
            v, e = _parse_decimal(valor_cell.value, HOJA, campo_cell.row,
                                  "monto_matricula_base")
            if e:
                errors.append(e)
            else:
                cfg["monto_matricula_base"] = v
        elif campo == "monto_pension_base":
            v, e = _parse_decimal(valor_cell.value, HOJA, campo_cell.row,
                                  "monto_pension_base")
            if e:
                errors.append(e)
            else:
                cfg["monto_pension_base"] = v
    # Defaults razonables si faltan
    if "anio_escolar" not in cfg:
        errors.append(_err(HOJA, "-", "anio_escolar: falta en la hoja 5_Configuracion"))
    if "monto_matricula_base" not in cfg:
        errors.append(_err(HOJA, "-", "monto_matricula_base: falta en la hoja 5_Configuracion"))
    if "monto_pension_base" not in cfg:
        errors.append(_err(HOJA, "-", "monto_pension_base: falta en la hoja 5_Configuracion"))
    return cfg


# ---------------------------------------------------------------------------
# Entry point del parsing
# ---------------------------------------------------------------------------

def parse_excel(file_obj):
    """Parsea el archivo Excel a un dict + lista de errores de formato."""
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:
        return {
            "aulas": [], "profesores": [], "alumnos": [], "apoderados": [], "config": {},
            "errors": [_err("(global)", "-", f"No se pudo abrir el archivo: {exc}")],
        }

    missing = [s for s in EXPECTED_SHEETS if s not in wb.sheetnames]
    if missing:
        return {
            "aulas": [], "profesores": [], "alumnos": [], "apoderados": [], "config": {},
            "errors": [_err("(global)", "-",
                            f"Faltan hojas obligatorias: {', '.join(missing)}")],
        }

    errors = []
    aulas = _parse_aulas(wb["1_Aulas"], errors)
    profesores = _parse_profesores(wb["2_Profesores"], errors)
    alumnos = _parse_alumnos(wb["3_Alumnos"], errors)
    apoderados = _parse_apoderados(wb["4_Apoderados"], errors)
    config = _parse_config(wb["5_Configuracion"], errors)
    return {
        "aulas": aulas, "profesores": profesores, "alumnos": alumnos,
        "apoderados": apoderados, "config": config, "errors": errors,
    }


# ---------------------------------------------------------------------------
# Validación cross-sheet
# ---------------------------------------------------------------------------

def validate_cross_sheet(data):
    """Reglas que requieren mirar más de una hoja a la vez."""
    errors = []

    aulas_por_nombre = {}
    for a in data["aulas"]:
        if a["nombre"] in aulas_por_nombre:
            errors.append(_err("1_Aulas", a["_fila"],
                               f"aula '{a['nombre']}' duplicada"))
        aulas_por_nombre[a["nombre"]] = a

    profesores_por_dni = {}
    for p in data["profesores"]:
        if p["dni"] in profesores_por_dni:
            errors.append(_err("2_Profesores", p["_fila"],
                               f"dni {p['dni']} duplicado"))
        profesores_por_dni[p["dni"]] = p
        if p["aula_titular"] and p["aula_titular"] not in aulas_por_nombre:
            errors.append(_err("2_Profesores", p["_fila"],
                               f"aula_titular '{p['aula_titular']}' no existe en 1_Aulas"))
        if p["aula_auxiliar"] and p["aula_auxiliar"] not in aulas_por_nombre:
            errors.append(_err("2_Profesores", p["_fila"],
                               f"aula_auxiliar '{p['aula_auxiliar']}' no existe en 1_Aulas"))
        if p["tipo"] == "TITULAR" and p["aula_auxiliar"]:
            errors.append(_err("2_Profesores", p["_fila"],
                               "profesor TITULAR no puede tener aula_auxiliar"))
        if p["tipo"] == "AUXILIAR" and p["aula_titular"]:
            errors.append(_err("2_Profesores", p["_fila"],
                               "profesor AUXILIAR no puede tener aula_titular"))

    # Solo 1 titular y 1 auxiliar por aula
    titulares_por_aula = defaultdict(list)
    auxiliares_por_aula = defaultdict(list)
    for p in data["profesores"]:
        if p["tipo"] == "TITULAR" and p["aula_titular"]:
            titulares_por_aula[p["aula_titular"]].append(p)
        if p["tipo"] == "AUXILIAR" and p["aula_auxiliar"]:
            auxiliares_por_aula[p["aula_auxiliar"]].append(p)
    for aula, profes in titulares_por_aula.items():
        if len(profes) > 1:
            for p in profes:
                errors.append(_err("2_Profesores", p["_fila"],
                                   f"aula '{aula}' tiene más de un profesor titular"))
    for aula, profes in auxiliares_por_aula.items():
        if len(profes) > 1:
            for p in profes:
                errors.append(_err("2_Profesores", p["_fila"],
                                   f"aula '{aula}' tiene más de un profesor auxiliar"))

    # Emails de profesores únicos (si vienen)
    emails_seen = {}
    for p in data["profesores"]:
        if p["email"]:
            if p["email"].lower() in emails_seen:
                errors.append(_err("2_Profesores", p["_fila"],
                                   f"email {p['email']} duplicado entre profesores"))
            emails_seen[p["email"].lower()] = p

    alumnos_por_dni = {}
    for a in data["alumnos"]:
        if a["dni"] in alumnos_por_dni:
            errors.append(_err("3_Alumnos", a["_fila"],
                               f"dni {a['dni']} duplicado"))
        alumnos_por_dni[a["dni"]] = a
        if a["aula"] not in aulas_por_nombre:
            errors.append(_err("3_Alumnos", a["_fila"],
                               f"aula '{a['aula']}' no existe en 1_Aulas"))

    # Apoderados: dni_alumno debe existir; al menos 1 principal por alumno;
    # máximo 2 apoderados por alumno; DNI único por alumno.
    apoderados_por_alumno = defaultdict(list)
    for ap in data["apoderados"]:
        if ap["dni_alumno"] not in alumnos_por_dni:
            errors.append(_err("4_Apoderados", ap["_fila"],
                               f"dni_alumno {ap['dni_alumno']} no existe en 3_Alumnos"))
            continue
        apoderados_por_alumno[ap["dni_alumno"]].append(ap)

    for dni_alumno, lista in apoderados_por_alumno.items():
        principales = [ap for ap in lista if ap["es_principal"]]
        if not principales:
            errors.append(_err("4_Apoderados", lista[0]["_fila"],
                               f"el alumno {dni_alumno} no tiene ningún apoderado marcado como principal"))
        elif len(principales) > 1:
            for ap in principales:
                errors.append(_err("4_Apoderados", ap["_fila"],
                                   f"el alumno {dni_alumno} tiene más de un apoderado principal"))
        if len(lista) > 2:
            for ap in lista:
                errors.append(_err("4_Apoderados", ap["_fila"],
                                   f"el alumno {dni_alumno} tiene más de 2 apoderados (máximo permitido: 2)"))
        dnis_apoderados = [ap["dni"] for ap in lista]
        if len(dnis_apoderados) != len(set(dnis_apoderados)):
            for ap in lista:
                errors.append(_err("4_Apoderados", ap["_fila"],
                                   f"DNIs de apoderado duplicados para el alumno {dni_alumno}"))

    # Cada alumno debe tener al menos 1 apoderado
    for dni_alumno, alumno in alumnos_por_dni.items():
        if dni_alumno not in apoderados_por_alumno:
            errors.append(_err("4_Apoderados", "-",
                               f"el alumno {dni_alumno} ({alumno['nombres']} {alumno['apellidos']}) no tiene ningún apoderado en 4_Apoderados"))

    return errors


# ---------------------------------------------------------------------------
# Sanity check: schema vacío
# ---------------------------------------------------------------------------

def is_schema_empty(tenant):
    """El import asume jardín vacío. Si ya hay datos, devolver False."""
    from apps.classrooms.models import Classroom
    from apps.students.models import Student
    from apps.teachers.models import Teacher

    with schema_context(tenant.schema_name):
        if (Classroom.objects.exists()
                or Student.objects.exists()
                or Teacher.objects.exists()):
            return False
    return True


# ---------------------------------------------------------------------------
# Ejecutor (creación real)
# ---------------------------------------------------------------------------

def _serializar(data):
    """Convierte el dict parseado a tipos JSON-serializables para session."""
    import copy
    out = copy.deepcopy(data)
    for a in out["alumnos"]:
        a["fecha_nacimiento"] = a["fecha_nacimiento"].isoformat()
        if a.get("monto_matricula") is not None:
            a["monto_matricula"] = str(a["monto_matricula"])
        if a.get("monto_pension") is not None:
            a["monto_pension"] = str(a["monto_pension"])
    for p in out["profesores"]:
        p["fecha_ingreso"] = p["fecha_ingreso"].isoformat()
    cfg = out["config"]
    if "monto_matricula_base" in cfg:
        cfg["monto_matricula_base"] = str(cfg["monto_matricula_base"])
    if "monto_pension_base" in cfg:
        cfg["monto_pension_base"] = str(cfg["monto_pension_base"])
    return out


def _deserializar(data):
    """Inversa de _serializar."""
    from datetime import date as _date
    for a in data["alumnos"]:
        if isinstance(a["fecha_nacimiento"], str):
            a["fecha_nacimiento"] = _date.fromisoformat(a["fecha_nacimiento"])
        if a.get("monto_matricula") is not None and isinstance(a["monto_matricula"], str):
            a["monto_matricula"] = Decimal(a["monto_matricula"])
        if a.get("monto_pension") is not None and isinstance(a["monto_pension"], str):
            a["monto_pension"] = Decimal(a["monto_pension"])
    for p in data["profesores"]:
        if isinstance(p["fecha_ingreso"], str):
            p["fecha_ingreso"] = _date.fromisoformat(p["fecha_ingreso"])
    cfg = data["config"]
    if isinstance(cfg.get("monto_matricula_base"), str):
        cfg["monto_matricula_base"] = Decimal(cfg["monto_matricula_base"])
    if isinstance(cfg.get("monto_pension_base"), str):
        cfg["monto_pension_base"] = Decimal(cfg["monto_pension_base"])
    return data


def execute_import(tenant, data, user=None):
    """
    Crea Aulas → Profesores → asigna a Aulas → Alumnos → Apoderados →
    Enrollment + MonthlyFee + 10 Payments por alumno.

    Todo en `schema_context(tenant.schema_name)` y `transaction.atomic()`.

    Returns:
        dict con counts por entidad y la lista de items creados.
    """
    from apps.classrooms.models import Classroom
    from apps.enrollments.models import Enrollment
    from apps.payments.models import MonthlyFee, Payment
    from apps.students.models import Guardian, Student
    from apps.teachers.models import Teacher

    data = _deserializar(data)
    cfg = data["config"]
    anio = cfg["anio_escolar"]
    matricula_base = cfg["monto_matricula_base"]
    pension_base = cfg["monto_pension_base"]

    counts = {
        "aulas": 0, "profesores": 0, "alumnos": 0, "apoderados": 0,
        "enrollments": 0, "monthly_fees": 0, "payments": 0,
    }

    with schema_context(tenant.schema_name):
        with transaction.atomic():
            # 1. Aulas
            aulas_creadas = {}
            for a in data["aulas"]:
                obj = Classroom.objects.create(
                    nombre=a["nombre"],
                    nivel_edad=a["nivel_edad"],
                    capacidad=a["capacidad"],
                )
                aulas_creadas[a["nombre"]] = obj
                counts["aulas"] += 1

            # 2. Profesores
            profes_por_dni = {}
            for p in data["profesores"]:
                obj = Teacher.objects.create(
                    dni=p["dni"], nombres=p["nombres"], apellidos=p["apellidos"],
                    tipo=p["tipo"], telefono=p["telefono"],
                    email=p["email"] or None,
                    especialidad=p["especialidad"] or "",
                    fecha_ingreso=p["fecha_ingreso"],
                )
                profes_por_dni[p["dni"]] = obj
                counts["profesores"] += 1

                # 3. Asignar profesor a aula
                if p["aula_titular"]:
                    aula = aulas_creadas.get(p["aula_titular"])
                    if aula:
                        aula.profesor_titular = obj
                        aula.save(update_fields=["profesor_titular"])
                if p["aula_auxiliar"]:
                    aula = aulas_creadas.get(p["aula_auxiliar"])
                    if aula:
                        aula.profesor_auxiliar = obj
                        aula.save(update_fields=["profesor_auxiliar"])

            # 4. Alumnos
            alumnos_por_dni = {}
            for a in data["alumnos"]:
                obj = Student.objects.create(
                    dni=a["dni"], nombres=a["nombres"], apellidos=a["apellidos"],
                    fecha_nacimiento=a["fecha_nacimiento"], genero=a["genero"],
                    classroom=aulas_creadas[a["aula"]],
                    estado=Student.Estado.ACTIVO,
                    fecha_ingreso=date.today(),
                )
                alumnos_por_dni[a["dni"]] = obj
                counts["alumnos"] += 1

                # 6. Enrollment (matrícula).
                # OJO con `or`: Decimal('0.00') es falsy en Python, así que
                # un override de "0.00" (beca completa) NO debe caer al base.
                # Usamos comparación explícita `is not None`.
                override_mat = a.get("monto_matricula")
                costo_matricula = override_mat if override_mat is not None else matricula_base
                Enrollment.objects.create(
                    student=obj, classroom=obj.classroom, anio_escolar=anio,
                    costo_matricula=costo_matricula, created_by=user,
                )
                counts["enrollments"] += 1

                # 7. MonthlyFee (pensión). Mismo cuidado con el override 0.00.
                override_pen = a.get("monto_pension")
                monto_pension = override_pen if override_pen is not None else pension_base
                mf = MonthlyFee.objects.create(
                    student=obj, anio_escolar=anio,
                    monto_mensual=monto_pension,
                    dia_vencimiento=DIA_VENCIMIENTO_DEFAULT,
                )
                counts["monthly_fees"] += 1

                # 8. 10 Payments (marzo-diciembre) en estado PENDIENTE
                for mes in range(MES_INICIO_PENSION, MES_FIN_PENSION + 1):
                    Payment.objects.create(
                        student=obj, monthly_fee=mf,
                        mes=mes, anio=anio, monto=monto_pension,
                        estado=Payment.Estado.PENDIENTE,
                        fecha_vencimiento=date(anio, mes, DIA_VENCIMIENTO_DEFAULT),
                    )
                    counts["payments"] += 1

            # 5. Apoderados
            for ap in data["apoderados"]:
                alumno = alumnos_por_dni.get(ap["dni_alumno"])
                if alumno is None:
                    continue  # validación ya descartó este caso
                Guardian.objects.create(
                    student=alumno, dni=ap["dni"], nombres=ap["nombres"],
                    apellidos=ap["apellidos"], telefono=ap["telefono"],
                    email=ap["email"] or None, parentesco=ap["parentesco"],
                    es_principal=ap["es_principal"],
                )
                counts["apoderados"] += 1

    return counts


# Reexportar helpers de session para que el admin no toque internals
serialize_for_session = _serializar
deserialize_from_session = _deserializar
