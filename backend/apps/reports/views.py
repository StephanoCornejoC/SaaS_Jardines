from datetime import date
from decimal import Decimal
from tempfile import SpooledTemporaryFile

from django.db.models import Count, Q
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from shared.validators import validate_month_param, validate_year_param

# Umbral para que el archivo se mantenga en RAM. Si el reporte crece más,
# SpooledTemporaryFile lo derrama automáticamente a disco temporal —
# el SO maneja la limpieza al cerrar el handle.
_SPOOL_MAX_BYTES = 4 * 1024 * 1024  # 4 MB


def _style_header(ws, headers, row=1):
    """Aplica estilos al encabezado de una hoja."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _build_response(wb, filename):
    """
    Genera FileResponse con el archivo Excel. Usa SpooledTemporaryFile que
    mantiene el archivo en RAM hasta 4MB y lo derrama a disco si crece más.
    FileResponse luego lo envía en chunks al cliente, evitando que el
    workbook entero permanezca duplicado en memoria mientras se serializa.
    """
    spool = SpooledTemporaryFile(max_size=_SPOOL_MAX_BYTES)
    wb.save(spool)
    spool.seek(0)
    return FileResponse(
        spool,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class ReportViewSet(viewsets.ViewSet):
    """
    ViewSet para generación de reportes Excel.
    """
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"], url_path="morosidad-excel")
    def morosidad_excel(self, request):
        """
        Genera reporte Excel de morosidad: pensiones cuyo vencimiento ya pasó
        y siguen pendientes (no pagadas ni exoneradas). Solo meses lectivos.
        """
        from apps.payments.models import Payment

        mes = validate_month_param(request.query_params.get("mes"))
        anio = validate_year_param(request.query_params.get("anio")) or date.today().year

        today = date.today()
        payments_qs = Payment.objects.filter(
            anio=anio,
            mes__gte=3,
            mes__lte=12,
            fecha_vencimiento__lt=today,
        ).exclude(
            estado__in=[Payment.Estado.PAGADO, Payment.Estado.EXONERADO],
        ).select_related("student", "student__classroom").order_by(
            "student__apellidos", "student__nombres", "mes",
        )

        if mes:
            payments_qs = payments_qs.filter(mes=mes)

        wb = Workbook()
        ws = wb.active
        ws.title = "Morosidad"

        headers = [
            "Alumno",
            "DNI",
            "Aula",
            "Mes",
            "Año",
            "Monto",
            "Fecha Vencimiento",
            "Días Vencido",
        ]
        _style_header(ws, headers)

        today = date.today()
        for row_idx, payment in enumerate(payments_qs, 2):
            dias_vencido = (today - payment.fecha_vencimiento).days
            ws.cell(row=row_idx, column=1, value=str(payment.student))
            ws.cell(row=row_idx, column=2, value=payment.student.dni)
            ws.cell(
                row=row_idx,
                column=3,
                value=str(payment.student.classroom) if payment.student.classroom else "Sin aula",
            )
            ws.cell(row=row_idx, column=4, value=payment.mes)
            ws.cell(row=row_idx, column=5, value=payment.anio)
            ws.cell(row=row_idx, column=6, value=payment.monto)
            ws.cell(
                row=row_idx,
                column=7,
                value=payment.fecha_vencimiento.strftime("%d/%m/%Y"),
            )
            ws.cell(row=row_idx, column=8, value=dias_vencido)

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

        filename = f"morosidad_{anio}.xlsx"
        return _build_response(wb, filename)

    @action(detail=False, methods=["get"], url_path="alumnos-excel")
    def alumnos_excel(self, request):
        """Genera reporte Excel con la lista de alumnos agrupada por aula."""
        from apps.students.models import Student

        estado = request.query_params.get("estado", "ACTIVO")

        students_qs = Student.objects.select_related("classroom")
        if estado:
            students_qs = students_qs.filter(estado=estado)
        students_qs = students_qs.order_by(
            "classroom__nivel_edad",
            "classroom__nombre",
            "apellidos",
            "nombres",
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Alumnos"

        headers = [
            "Apellidos",
            "Nombres",
            "DNI",
            "Fecha Nacimiento",
            "Edad",
            "Género",
            "Estado",
            "Fecha Ingreso",
        ]

        # Agrupar por aula. Cada grupo: header de aula + headers + filas + fila vacía.
        from collections import OrderedDict

        groups = OrderedDict()
        for s in students_qs:
            key = (
                s.classroom.id if s.classroom else None,
                str(s.classroom) if s.classroom else "Sin aula asignada",
            )
            groups.setdefault(key, []).append(s)

        group_font = Font(bold=True, size=12, color="0F766E")
        group_fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")

        row_idx = 1
        for (_aula_id, aula_label), alumnos in groups.items():
            # Título de aula
            ws.cell(row=row_idx, column=1, value=f"Aula: {aula_label}  ({len(alumnos)} alumno(s))")
            ws.cell(row=row_idx, column=1).font = group_font
            ws.cell(row=row_idx, column=1).fill = group_fill
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            row_idx += 1

            # Encabezado por grupo
            _style_header(ws, headers, row=row_idx)
            row_idx += 1

            for s in alumnos:
                ws.cell(row=row_idx, column=1, value=s.apellidos)
                ws.cell(row=row_idx, column=2, value=s.nombres)
                ws.cell(row=row_idx, column=3, value=s.dni)
                ws.cell(row=row_idx, column=4, value=s.fecha_nacimiento.strftime("%d/%m/%Y"))
                ws.cell(row=row_idx, column=5, value=s.edad)
                ws.cell(row=row_idx, column=6, value=s.get_genero_display())
                ws.cell(row=row_idx, column=7, value=s.get_estado_display())
                ws.cell(row=row_idx, column=8, value=s.fecha_ingreso.strftime("%d/%m/%Y"))
                row_idx += 1

            # Fila en blanco entre grupos
            row_idx += 1

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

        filename = f"alumnos_{estado.lower()}.xlsx"
        return _build_response(wb, filename)

    @action(detail=False, methods=["get"], url_path="asistencia-excel")
    def asistencia_excel(self, request):
        """
        Genera reporte Excel de asistencia. Si no se pasa classroom_id, incluye
        todas las aulas (cada una en una hoja separada).
        """
        from apps.attendance.models import Attendance
        from apps.classrooms.models import Classroom

        classroom_id = request.query_params.get("classroom_id")
        mes = validate_month_param(request.query_params.get("mes")) or date.today().month
        anio = validate_year_param(request.query_params.get("anio")) or date.today().year

        if classroom_id:
            try:
                classrooms = [Classroom.objects.get(pk=int(classroom_id))]
            except Classroom.DoesNotExist:
                return Response(
                    {"error": "Aula no encontrada."},
                    status=status.HTTP_404_NOT_FOUND,
                )
        else:
            classrooms = list(Classroom.objects.order_by("nivel_edad", "nombre"))

        if not classrooms:
            return Response(
                {"error": "No hay aulas registradas."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = Workbook()
        wb.remove(wb.active)  # quitar la hoja por defecto

        for classroom in classrooms:
            attendances = Attendance.objects.filter(
                classroom=classroom,
                fecha__month=mes,
                fecha__year=anio,
            ).select_related("student")

            report = (
                attendances.values("student__apellidos", "student__nombres")
                .annotate(
                    total_presentes=Count("id", filter=Q(estado=Attendance.Estado.PRESENTE)),
                    total_ausentes=Count("id", filter=Q(estado=Attendance.Estado.AUSENTE)),
                    total_tardanzas=Count("id", filter=Q(estado=Attendance.Estado.TARDANZA)),
                    total_justificados=Count(
                        "id", filter=Q(estado=Attendance.Estado.JUSTIFICADO)
                    ),
                    total_dias=Count("id"),
                )
                .order_by("student__apellidos", "student__nombres")
            )

            sheet_name = classroom.nombre[:31]  # Excel limita a 31 chars
            ws = wb.create_sheet(title=sheet_name)
            self._render_attendance_sheet(ws, classroom, mes, anio, report)

        if not wb.sheetnames:
            wb.create_sheet(title="Sin datos")

        suffix = classrooms[0].nombre if len(classrooms) == 1 else "todas"
        filename = f"asistencia_{suffix}_{mes}_{anio}.xlsx"
        return _build_response(wb, filename)

    def _render_attendance_sheet(self, ws, classroom, mes, anio, report):
        ws.merge_cells("A1:H1")
        title_cell = ws.cell(
            row=1, column=1,
            value=f"Reporte de Asistencia - {classroom.nombre} - {mes}/{anio}",
        )
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "Alumno",
            "Presentes",
            "Ausentes",
            "Tardanzas",
            "Justificados",
            "Total Días",
            "% Asistencia",
        ]
        _style_header(ws, headers, row=3)

        for row_idx, row in enumerate(report, 4):
            nombre = f"{row['student__apellidos']}, {row['student__nombres']}"
            total = row["total_dias"]
            porcentaje = (row["total_presentes"] / total * 100) if total > 0 else 0

            ws.cell(row=row_idx, column=1, value=nombre)
            ws.cell(row=row_idx, column=2, value=row["total_presentes"])
            ws.cell(row=row_idx, column=3, value=row["total_ausentes"])
            ws.cell(row=row_idx, column=4, value=row["total_tardanzas"])
            ws.cell(row=row_idx, column=5, value=row["total_justificados"])
            ws.cell(row=row_idx, column=6, value=total)
            ws.cell(row=row_idx, column=7, value=f"{porcentaje:.1f}%")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

    @action(detail=False, methods=["get"], url_path="cashflow-excel")
    def cashflow_excel(self, request):
        """Genera reporte Excel de ingresos y egresos."""
        from apps.cashflow.models import CashTransaction

        mes = validate_month_param(request.query_params.get("mes"))
        anio = validate_year_param(request.query_params.get("anio")) or date.today().year

        transactions_qs = CashTransaction.objects.filter(
            fecha__year=anio
        ).select_related("categoria")

        if mes:
            transactions_qs = transactions_qs.filter(fecha__month=mes)

        wb = Workbook()
        ws = wb.active
        ws.title = "Flujo de Caja"

        # Titulo
        periodo = f"{mes}/{anio}" if mes else str(anio)
        ws.merge_cells("A1:G1")
        title_cell = ws.cell(
            row=1, column=1, value=f"Reporte de Flujo de Caja - {periodo}"
        )
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "Fecha",
            "Tipo",
            "Categoría",
            "Descripción",
            "Monto",
        ]
        _style_header(ws, headers, row=3)

        total_ingresos = Decimal("0")
        total_egresos = Decimal("0")

        for row_idx, tx in enumerate(transactions_qs.order_by("fecha"), 4):
            ws.cell(row=row_idx, column=1, value=tx.fecha.strftime("%d/%m/%Y"))
            ws.cell(row=row_idx, column=2, value=tx.get_tipo_display())
            ws.cell(row=row_idx, column=3, value=str(tx.categoria))
            ws.cell(row=row_idx, column=4, value=tx.descripcion)
            ws.cell(row=row_idx, column=5, value=tx.monto)

            if tx.tipo == CashTransaction.Tipo.INGRESO:
                total_ingresos += tx.monto
            else:
                total_egresos += tx.monto

        # Resumen al final
        last_row = transactions_qs.count() + 5
        summary_font = Font(bold=True, size=11)

        ws.cell(row=last_row, column=3, value="Total Ingresos:").font = summary_font
        ws.cell(row=last_row, column=5, value=total_ingresos).font = summary_font

        ws.cell(row=last_row + 1, column=3, value="Total Egresos:").font = summary_font
        ws.cell(row=last_row + 1, column=5, value=total_egresos).font = summary_font

        ws.cell(row=last_row + 2, column=3, value="Balance:").font = summary_font
        balance_cell = ws.cell(
            row=last_row + 2, column=5, value=total_ingresos - total_egresos
        )
        balance_cell.font = summary_font

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

        filename = f"flujo_caja_{periodo.replace('/', '_')}.xlsx"
        return _build_response(wb, filename)
